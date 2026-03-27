"""
Excel 转 HTML 工具
使用 openpyxl (xlsx) 和 xlrd (xls) 读取 Excel 文件，生成 HTML 表格预览
"""
import io
import logging
import tempfile
import os
from datetime import datetime

logger = logging.getLogger(__name__)

DEFAULT_MAX_ROWS = 1000
DEFAULT_MAX_COLS = 100


def _should_try_xlrd(error_message):
    """判断 openpyxl 失败后是否应回退到 xlrd（主要针对 .xls）"""
    if not error_message:
        return False

    keywords = [
        'openpyxl does not support the old .xls',
        'file is not a zip file',
        'file contains no valid workbook part',
        'invalid file'
    ]
    lower_msg = error_message.lower()
    return any(keyword in lower_msg for keyword in keywords)


def excel_to_html(excel_blob, max_rows=DEFAULT_MAX_ROWS, max_cols=DEFAULT_MAX_COLS):
    """
    将 Excel 文件转换为 HTML

    Args:
        excel_blob: Excel 文件的二进制数据

    Returns:
        dict: {
            'html': str,  # HTML 内容
            'sheets': list,  # 工作表列表
            'active_sheet': int  # 活动工作表索引
        }
        如果转换失败返回 None
    """
    temp_file = None
    try:
        # 旧版 xls（OLE2）优先走 xlrd
        if excel_blob[:4] == b'\xD0\xCF\x11\xE0':
            logger.info('检测到 OLE2 格式，优先使用 xlrd 处理 .xls')
            return _excel_to_html_xlrd(excel_blob, max_rows=max_rows, max_cols=max_cols)

        # 先尝试用 openpyxl 打开（支持 xlsx/xlsm）
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Alignment, Color

            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.write(excel_blob)
            temp_file.close()

            workbook = openpyxl.load_workbook(temp_file.name, data_only=True)
            sheets = workbook.worksheets
            sheet_list = [{'name': sheet.title, 'index': i} for i, sheet in enumerate(sheets)]
            html_content = _generate_workbook_html_openpyxl(workbook, sheets, max_rows=max_rows, max_cols=max_cols)

            return {
                'html': html_content,
                'sheets': sheet_list,
                'active_sheet': 0
            }

        except Exception as e:
            # openpyxl 失败，尝试用 xlrd（支持 xls 旧格式）
            if _should_try_xlrd(str(e)):
                logger.info('openpyxl 无法打开，尝试使用 xlrd')
                return _excel_to_html_xlrd(excel_blob, max_rows=max_rows, max_cols=max_cols)
            else:
                raise

    except Exception as e:
        logger.error(f'Excel 转 HTML 失败: {str(e)}', exc_info=True)
        return None
    finally:
        # 清理临时文件
        if temp_file and os.path.exists(temp_file.name):
            try:
                os.remove(temp_file.name)
            except:
                pass


def _excel_to_html_xlrd(excel_blob, max_rows=DEFAULT_MAX_ROWS, max_cols=DEFAULT_MAX_COLS):
    """使用 xlrd 处理 xls 旧格式"""
    try:
        import xlrd
    except ImportError:
        logger.error('未安装 xlrd，请运行: pip install xlrd')
        return None

    temp_file = None
    try:
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xls')
        temp_file.write(excel_blob)
        temp_file.close()

        # 打开工作簿
        workbook = xlrd.open_workbook(temp_file.name)
        sheets = workbook.sheets()
        sheet_list = [{'name': sheet.name, 'index': i} for i, sheet in enumerate(sheets)]

        # 生成 HTML
        html_content = _generate_workbook_html_xlrd(workbook, sheets, max_rows=max_rows, max_cols=max_cols)

        return {
            'html': html_content,
            'sheets': sheet_list,
            'active_sheet': 0
        }

    finally:
        if temp_file and os.path.exists(temp_file.name):
            try:
                os.remove(temp_file.name)
            except:
                pass


def _generate_workbook_html_openpyxl(workbook, sheets, max_rows=DEFAULT_MAX_ROWS, max_cols=DEFAULT_MAX_COLS):
    """生成整个工作簿的 HTML (openpyxl 版本)"""
    return _generate_workbook_html_base(
        sheets,
        lambda sheet, row, col: _get_cell_value_openpyxl(sheet.cell(row + 1, col + 1)),
        lambda sheet: sheet.max_row,
        lambda sheet: sheet.max_column,
        lambda sheet, row, col: _get_cell_style_openpyxl(sheet.cell(row + 1, col + 1)),
        lambda sheet: sheet.title,
        lambda sheet, row, col: sheet.cell(row + 1, col + 1),
        lambda sheet: _get_openpyxl_used_bounds(sheet),
        max_rows=max_rows,
        max_cols=max_cols
    )


def _generate_workbook_html_xlrd(workbook, sheets, max_rows=DEFAULT_MAX_ROWS, max_cols=DEFAULT_MAX_COLS):
    """生成整个工作簿的 HTML (xlrd 版本)"""
    return _generate_workbook_html_base(
        sheets,
        lambda sheet, row, col: _get_cell_value_xlrd(sheet, row, col),
        lambda sheet: sheet.nrows,
        lambda sheet: sheet.ncols,
        lambda sheet, row, col: '',  # xlrd 不支持样式
        lambda sheet: sheet.name,
        lambda sheet, row, col: None,  # xlrd 单元格对象
        None,
        max_rows=max_rows,
        max_cols=max_cols
    )


def _generate_workbook_html_base(
    sheets,
    get_value,
    get_nrows,
    get_ncols,
    get_style,
    get_name,
    get_cell,
    get_bounds=None,
    max_rows=DEFAULT_MAX_ROWS,
    max_cols=DEFAULT_MAX_COLS
):
    """生成整个工作簿的 HTML (通用版本)"""
    html_parts = []

    # HTML 头部
    html_parts.append('''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }

        html, body {
            height: 100%;
            width: 100%;
        }

        body {
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            background: #f5f5f5;
            padding: 8px;
            overflow: hidden;
        }

        .excel-container {
            height: 100%;
            background: white;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            overflow: hidden;
            display: flex;
            flex-direction: column;
        }

        .sheet-tabs {
            display: flex;
            background: #f0f0f0;
            border-bottom: 1px solid #ddd;
            overflow-x: auto;
            flex-shrink: 0;
        }

        .sheet-tab {
            padding: 10px 20px;
            cursor: pointer;
            border-right: 1px solid #ddd;
            background: #e0e0e0;
            transition: background 0.2s;
            white-space: nowrap;
            user-select: none;
        }

        .sheet-tab:hover {
            background: #d0d0d0;
        }

        .sheet-tab.active {
            background: white;
            border-bottom: 2px solid #2196F3;
            font-weight: 500;
        }

        .sheet-content {
            display: none;
            flex: 1;
            min-height: 0;
            overflow: auto;
        }

        .sheet-content.active {
            display: block;
            height: 100%;
        }

        table {
            border-collapse: collapse;
            width: 100%;
        }

        td {
            border: 1px solid #d0d0d0;
            padding: 4px 8px;
            white-space: pre;
            overflow: hidden;
            text-overflow: ellipsis;
            min-width: 50px;
            min-height: 20px;
        }

        td.header {
            background: #f5f5f5;
            font-weight: 500;
            text-align: center;
            color: #333;
        }
    </style>
</head>
<body>
    <div class="excel-container">
        <div class="sheet-tabs">
''')

    # 工作表标签
    for i, sheet in enumerate(sheets):
        active_class = 'active' if i == 0 else ''
        sheet_name_html = get_name(sheet).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')
        html_parts.append(f'            <div class="sheet-tab {active_class}" data-sheet="{i}">{sheet_name_html}</div>')

    html_parts.append('        </div>')

    # 各个工作表的内容
    for i, sheet in enumerate(sheets):
        active_class = 'active' if i == 0 else ''
        html_parts.append(f'        <div class="sheet-content {active_class}" data-sheet="{i}">')
        html_parts.append(
            _generate_sheet_html_base(
                sheet,
                get_value,
                get_nrows,
                get_ncols,
                get_style,
                get_cell,
                get_bounds=get_bounds,
                max_rows=max_rows,
                max_cols=max_cols
            )
        )
        html_parts.append('        </div>')

    # JavaScript
    html_parts.append('''
    </div>

    <script>
        // 工作表切换
        document.querySelectorAll('.sheet-tab').forEach(tab => {
            tab.addEventListener('click', function() {
                const sheetIndex = this.dataset.sheet;

                // 移除所有 active 类
                document.querySelectorAll('.sheet-tab').forEach(t => t.classList.remove('active'));
                document.querySelectorAll('.sheet-content').forEach(c => c.classList.remove('active'));

                // 添加 active 类到当前选中的工作表
                this.classList.add('active');
                document.querySelector(`.sheet-content[data-sheet="${sheetIndex}"]`).classList.add('active');
            });
        });
    </script>
</body>
</html>''')

    return '\n'.join(html_parts)


def _generate_sheet_html_base(
    sheet,
    get_value,
    get_nrows,
    get_ncols,
    get_style,
    get_cell,
    get_bounds=None,
    max_rows=DEFAULT_MAX_ROWS,
    max_cols=DEFAULT_MAX_COLS
):
    """生成单个工作表的 HTML (通用版本)"""
    html_parts = ['        <table>']

    # 获取工作表的实际使用范围
    max_row = get_nrows(sheet)
    max_col = get_ncols(sheet)

    # 默认渲染整张表（1-based）
    start_row_1 = 1
    start_col_1 = 1
    end_row_1 = max_row
    end_col_1 = max_col

    # 若可用，优先裁剪到“有值区域”，避免数据不在 A1 时出现大片空白
    if get_bounds:
        try:
            s_row, e_row, s_col, e_col = get_bounds(sheet)
            start_row_1 = max(1, s_row)
            start_col_1 = max(1, s_col)
            end_row_1 = max(start_row_1, e_row)
            end_col_1 = max(start_col_1, e_col)
        except Exception:
            pass

    # 限制最大行数和列数，避免过大的文件

    total_rows = max(0, end_row_1 - start_row_1 + 1)
    total_cols = max(0, end_col_1 - start_col_1 + 1)
    display_rows = min(total_rows, max_rows)
    display_cols = min(total_cols, max_cols)

    start_row_0 = start_row_1 - 1
    start_col_0 = start_col_1 - 1

    for row_idx in range(display_rows):
        html_parts.append('            <tr>')

        for col_idx in range(display_cols):
            real_row_0 = start_row_0 + row_idx
            real_col_0 = start_col_0 + col_idx
            cell_value = get_value(sheet, real_row_0, real_col_0)
            style = get_style(sheet, real_row_0, real_col_0)

            # 第一行作为表头
            is_header = row_idx == 0
            header_class = ' header' if is_header else ''

            # 转义 HTML 特殊字符
            escaped_value = _escape_html(cell_value)

            html_parts.append(f'                <td class="cell{header_class}" style="{style}">{escaped_value}</td>')

        html_parts.append('            </tr>')

    html_parts.append('        </table>')

    if total_rows > max_rows or total_cols > max_cols:
        html_parts.append(f'        <div style="padding: 10px; background: #fff3cd; color: #856404; text-align: center;">')
        html_parts.append(f'            ⚠️ 仅显示前 {max_rows} 行 × {max_cols} 列，可在预览窗口切换完整模式')
        html_parts.append('        </div>')

    return '\n'.join(html_parts)


def _get_cell_value_openpyxl(cell):
    """获取单元格值 (openpyxl 版本)"""
    value = cell.value

    if value is None:
        return ''

    # 处理日期时间
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')

    # 处理数字
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        if abs(value) < 0.01 or abs(value) >= 1000000:
            return f'{value:.6g}'
        return str(value)

    if isinstance(value, int):
        return str(value)

    return str(value)


def _get_openpyxl_used_bounds(sheet):
    """
    返回 openpyxl sheet 的有效区域边界 (start_row, end_row, start_col, end_col), 全部 1-based。
    若无法判定，回退到 (1, max_row, 1, max_col)。
    """
    try:
        dim = sheet.calculate_dimension()  # 例如 A1:D20 或 A1
        if ':' in dim:
            start_ref, end_ref = dim.split(':', 1)
        else:
            start_ref = end_ref = dim

        from openpyxl.utils.cell import coordinate_to_tuple
        s_row, s_col = coordinate_to_tuple(start_ref)
        e_row, e_col = coordinate_to_tuple(end_ref)

        # 若整表为空，openpyxl 常返回 A1:A1，这里按至少 1x1 处理
        return s_row, e_row, s_col, e_col
    except Exception:
        return 1, max(1, sheet.max_row), 1, max(1, sheet.max_column)


def _get_cell_value_xlrd(sheet, row, col):
    """获取单元格值 (xlrd 版本)"""
    try:
        import xlrd
        cell = sheet.cell(row, col)
        value = cell.value

        if value is None:
            return ''

        # xlrd 的日期是数字类型
        if cell.ctype == 3:  # XL_CELL_DATE
            try:
                datetuple = xlrd.xldate_as_tuple(value, sheet.book.datemode)
                return datetime(*datetuple).strftime('%Y-%m-%d %H:%M:%S')
            except:
                return str(value)

        if cell.ctype == 2:  # XL_CELL_NUMBER
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value)

        return str(value)
    except:
        return ''


def _get_cell_style_openpyxl(cell):
    """获取单元格样式 (openpyxl 版本)"""
    styles = []

    # 背景色
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        rgb = _get_color_string(cell.fill.fgColor.rgb)
        if rgb and rgb != '00000000':
            rgb = rgb[2:] if len(rgb) > 6 else rgb
            styles.append(f'background-color: #{rgb}')

    # 字体样式
    if cell.font:
        if cell.font.bold:
            styles.append('font-weight: bold')
        if cell.font.italic:
            styles.append('font-style: italic')
        if cell.font.color and cell.font.color.rgb:
            rgb = _get_color_string(cell.font.color.rgb)
            if rgb:
                rgb = rgb[2:] if len(rgb) > 6 else rgb
                if rgb != '000000':
                    styles.append(f'color: #{rgb}')
        if cell.font.size:
            styles.append(f'font-size: {cell.font.size}pt')

    # 对齐方式
    if cell.alignment:
        if cell.alignment.horizontal:
            styles.append(f'text-align: {cell.alignment.horizontal}')
        if cell.alignment.vertical:
            styles.append(f'vertical-align: {cell.alignment.vertical}')
        if cell.alignment.wrap_text:
            styles.append('white-space: normal')

    return '; '.join(styles)


def _get_color_string(color):
    """获取颜色字符串，处理 RGB 对象"""
    if color is None:
        return None

    if isinstance(color, str):
        return color

    try:
        if hasattr(color, 'rgb'):
            return str(color.rgb)
        return str(color)
    except:
        return None


def _escape_html(text):
    """转义 HTML 特殊字符"""
    if not text:
        return ''
    text = str(text)
    text = text.replace('&', '&amp;')
    text = text.replace('<', '&lt;')
    text = text.replace('>', '&gt;')
    text = text.replace('"', '&quot;')
    text = text.replace("'", '&#x27;')
    return text


def check_excel_to_html_available():
    """
    检查 Excel 转 HTML 是否可用

    Returns:
        bool: 是否可用
    """
    try:
        import openpyxl
        return True
    except ImportError:
        return False
